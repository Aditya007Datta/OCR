"""Install:
  pip install ddgs requests pdfplumber python-docx openai pandas openpyxl rapidfuzz
"""

import os
import re
import json
import logging
import hashlib
import mimetypes
from pathlib import Path
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional
from urllib.parse import urlparse

import requests
import pdfplumber
import docx as python_docx
from rapidfuzz import fuzz
from ddgs import DDGS
from openai import AzureOpenAI


# ─────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
)
log = logging.getLogger("standards_pipeline")


# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────

def _require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise EnvironmentError(
            f"Required environment variable '{name}' is not set."
        )
    return value


AZURE_ENDPOINT    = _require_env("AZURE_OPENAI_ENDPOINT")
AZURE_KEY         = _require_env("AZURE_OPENAI_KEY")
AZURE_DEPLOYMENT  = _require_env("AZURE_OPENAI_DEPLOYMENT")
AZURE_API_VERSION = "2024-02-01"

DOWNLOAD_DIR = Path(os.getenv("DOWNLOAD_DIR", "downloads"))
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "output"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

MAX_SEARCH_RESULTS    = int(os.getenv("MAX_SEARCH_RESULTS",    "15"))
MAX_CANDIDATES_TO_LLM = int(os.getenv("MAX_CANDIDATES_TO_LLM", "8"))
MAX_WORKERS           = int(os.getenv("MAX_WORKERS",           "6"))
REQUEST_TIMEOUT       = int(os.getenv("REQUEST_TIMEOUT",       "30"))
CHUNK_SIZE            = int(os.getenv("CHUNK_SIZE",            "350"))   # words
CHUNK_OVERLAP         = int(os.getenv("CHUNK_OVERLAP",         "60"))    # words
DEDUP_THRESHOLD       = int(os.getenv("DEDUP_THRESHOLD",       "88"))    # %

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".html", ".htm", ".txt"}


# ─────────────────────────────────────────────────────────────
# CLIENT
# ─────────────────────────────────────────────────────────────

azure_client = AzureOpenAI(
    api_key=AZURE_KEY,
    azure_endpoint=AZURE_ENDPOINT,
    api_version=AZURE_API_VERSION,
)


# ─────────────────────────────────────────────────────────────
# DATA MODELS
# ─────────────────────────────────────────────────────────────

@dataclass
class Candidate:
    url:     str
    title:   str
    snippet: str
    score:   float = 0.0
    preview: str   = ""


@dataclass
class TextChunk:
    text:            str
    section_heading: str   # e.g. "5.1 Access Control Policy"
    chunk_index:     int
    page_number:     int = 0


@dataclass
class Requirement:
    source_name:      str
    section_number:   str
    section_heading:  str
    topic:            str
    sub_topic:        str
    obligation_type:  str   # mandatory / recommended / conditional / negative / responsibility
    condition:        str   # e.g. "if encryption is used" — empty if unconditional
    actor:            str   # who must do this — e.g. "The organization", "System administrators"
    requirement:      str   # the obligation statement
    rationale:        str   # why this matters (auditor inference)


@dataclass
class PipelineResult:
    standard:     str
    requirements: list[Requirement] = field(default_factory=list)
    chosen_url:   Optional[str]     = None
    local_path:   Optional[Path]    = None
    error:        Optional[str]     = None


# ─────────────────────────────────────────────────────────────
# SEARCH + CANDIDATE SELECTION  (unchanged from previous version)
# ─────────────────────────────────────────────────────────────

_FILETYPE_TEMPLATES = [
    '"{standard}" official standard filetype:{fmt}',
    '"{standard}" specification filetype:{fmt}',
    '"{standard}" framework filetype:{fmt}',
]
_FALLBACK_TEMPLATES = [
    '"{standard}" site:iso.org OR site:nist.gov OR site:ietf.org OR site:ieee.org',
    '"{standard}" official requirements document PDF',
    '"{standard}" compliance standard official publication',
]
_PRIORITY_FORMATS = ["pdf", "docx"]

_AUTHORITATIVE_DOMAINS = [
    "iso.org", "nist.gov", "ietf.org", "w3.org", "ieee.org",
    "cis.org", "aicpa.org", "isaca.org", "pcisecuritystandards.org",
    "enisa.europa.eu", "bsigroup.com",
]
_JUNK_SIGNALS = [
    "blog", "tutorial", "quickstart", "getting-started", "how-to",
    "guide", "cheatsheet", "explainer", "overview", "what-is",
]

_URL_RE = re.compile(r"https?://[^\s\"'<>)\]]+")


def generate_queries(standard: str) -> list[str]:
    seen: set[str] = set()
    queries: list[str] = []

    def add(q: str) -> None:
        q = q.strip()
        if q not in seen:
            seen.add(q)
            queries.append(q)

    for fmt in _PRIORITY_FORMATS:
        for tmpl in _FILETYPE_TEMPLATES:
            add(tmpl.format(standard=standard, fmt=fmt))
    for tmpl in _FALLBACK_TEMPLATES:
        add(tmpl.format(standard=standard))
    return queries


def search_web(queries: list[str]) -> list[Candidate]:
    seen_urls: set[str] = set()
    candidates: list[Candidate] = []

    with DDGS() as ddgs:
        for query in queries:
            try:
                for r in ddgs.text(query, max_results=MAX_SEARCH_RESULTS):
                    url = r.get("href", "").strip()
                    if not url or url in seen_urls:
                        continue
                    seen_urls.add(url)
                    candidates.append(Candidate(
                        url=url,
                        title=r.get("title", ""),
                        snippet=r.get("body", ""),
                        score=1.0 / (len(candidates) + 1),
                    ))
            except Exception:
                log.warning("Search failed: %s", query, exc_info=True)

    log.info("Found %d unique candidates", len(candidates))
    return candidates


def score_candidates(candidates: list[Candidate], standard: str) -> list[Candidate]:
    slug1 = standard.lower().replace(" ", "-")
    slug2 = standard.lower().replace(" ", "_")

    for c in candidates:
        domain = urlparse(c.url).netloc.lower()
        path   = urlparse(c.url).path.lower()
        text   = (c.title + " " + c.snippet + " " + c.url).lower()

        if any(a in domain for a in _AUTHORITATIVE_DOMAINS): c.score += 2.5
        if c.url.lower().endswith(".pdf"):                    c.score += 1.2
        elif c.url.lower().endswith((".docx", ".doc")):       c.score += 0.6
        if slug1 in path or slug2 in path:                   c.score += 0.9
        if any(j in text for j in _JUNK_SIGNALS):            c.score -= 1.2

    return sorted(candidates, key=lambda c: c.score, reverse=True)


def fetch_preview(url: str, max_bytes: int = 6000) -> str:
    try:
        resp = requests.get(url, timeout=10, stream=True,
                            headers={"Range": f"bytes=0-{max_bytes}",
                                     "User-Agent": "StandardsPipeline/1.0"})
        resp.raise_for_status()
        ct = resp.headers.get("Content-Type", "").lower()
        raw = b""
        for chunk in resp.iter_content(1024):
            raw += chunk
            if len(raw) >= max_bytes:
                break
        if "pdf" in ct or url.lower().endswith(".pdf"):
            runs = re.findall(rb"[A-Za-z0-9 ,.\-:;()\[\]{}/\n]{6,}", raw)
            return b" ".join(runs).decode("ascii", errors="ignore")[:max_bytes]
        return raw.decode("utf-8", errors="ignore")[:max_bytes]
    except Exception:
        log.warning("Preview failed: %s", url, exc_info=True)
        return ""


def fetch_previews_parallel(candidates: list[Candidate]) -> list[Candidate]:
    top = candidates[:MAX_CANDIDATES_TO_LLM]
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(fetch_preview, c.url): c for c in top}
        for fut in as_completed(futures):
            futures[fut].preview = fut.result()
    return candidates


def choose_document_llm(standard: str, candidates: list[Candidate]) -> Optional[str]:
    if not candidates:
        return None

    block = "\n\n".join(
        f"[{i+1}] URL: {c.url}\n     Title: {c.title}\n"
        f"     Preview: {(c.preview or c.snippet)[:500]}"
        for i, c in enumerate(candidates[:MAX_CANDIDATES_TO_LLM])
    )
    prompt = (
        f'Standard: "{standard}"\n\nCandidates:\n{block}\n\n'
        "Respond with ONLY the exact URL of the best canonical document."
    )
    try:
        resp = azure_client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[
                {"role": "system", "content":
                    "You identify the single most authoritative official source document "
                    "for a compliance or technical standard."},
                {"role": "user", "content": prompt},
            ],
            temperature=0, max_tokens=200,
        )
        raw = resp.choices[0].message.content.strip()
        match = _URL_RE.search(raw)
        if match:
            url = match.group(0).rstrip(".,)")
            if any(c.url == url for c in candidates):
                return url
        for c in candidates:
            if c.url in raw:
                return c.url
        return None
    except Exception:
        log.error("LLM selection failed", exc_info=True)
        return None


# ─────────────────────────────────────────────────────────────
# DOWNLOAD
# ─────────────────────────────────────────────────────────────

def _safe_filename(url: str, content_type: str) -> str:
    path = urlparse(url).path
    name = Path(path).name
    ext  = Path(name).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        guessed = mimetypes.guess_extension(content_type.split(";")[0].strip())
        ext  = guessed if guessed and guessed in SUPPORTED_EXTENSIONS else ".bin"
        name = hashlib.md5(url.encode()).hexdigest()[:12] + ext
    return re.sub(r"[^\w.\-]", "_", name)


def download_file(url: str) -> Optional[Path]:
    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT, stream=True,
                            headers={"User-Agent": "StandardsPipeline/1.0"})
        resp.raise_for_status()
        ct       = resp.headers.get("Content-Type", "application/octet-stream")
        filepath = DOWNLOAD_DIR / _safe_filename(url, ct)
        with open(filepath, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)
        log.info("Downloaded → %s (%.1f KB)", filepath, filepath.stat().st_size / 1024)
        return filepath
    except Exception:
        log.error("Download failed: %s", url, exc_info=True)
        return None


# ─────────────────────────────────────────────────────────────
# TEXT EXTRACTION  (format-aware)
# ─────────────────────────────────────────────────────────────

# Heading patterns common in standards documents
_HEADING_RE = re.compile(
    r"^(\d+(\.\d+){0,3}[\s\t]+[A-Z][^\n]{3,80}|"   # "5.1.2 Access Control"
    r"Annex\s+[A-Z][\s\S]{0,60}|"                   # "Annex A — ..."
    r"[A-Z][A-Z\s]{4,50}(?:\n|$))",                  # "ACCESS CONTROL"
    re.MULTILINE,
)


def extract_text_with_structure(path: Path) -> list[tuple[str, str]]:
    """
    Returns list of (section_heading, paragraph_text) tuples.
    Preserves document structure so every chunk knows its section.
    """
    ext = path.suffix.lower()

    try:
        if ext == ".pdf":
            return _extract_pdf_structured(path)
        if ext in {".docx", ".doc"}:
            return _extract_docx_structured(path)
        if ext in {".html", ".htm"}:
            raw  = path.read_text(errors="ignore")
            text = re.sub(r"<[^>]+>", " ", raw)
            text = re.sub(r"\s+", " ", text).strip()
            return _split_by_headings(text)
        text = path.read_text(errors="ignore")
        return _split_by_headings(text)

    except Exception:
        log.error("Text extraction failed for %s", path, exc_info=True)
        return []


def _extract_pdf_structured(path: Path) -> list[tuple[str, str]]:
    """Extract PDF text page by page, detect headings by font size heuristic."""
    sections: list[tuple[str, str]] = []
    current_heading = "Introduction"
    current_lines:   list[str] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            # Try to get words with size info for heading detection
            words = page.extract_words(extra_attrs=["size"])
            if not words:
                continue

            # Find the dominant body font size (most common)
            sizes = [w.get("size", 12) for w in words if w.get("size")]
            if sizes:
                sizes_sorted = sorted(sizes)
                body_size = sizes_sorted[len(sizes_sorted) // 2]  # median
            else:
                body_size = 12

            page_text = page.extract_text() or ""
            for line in page_text.split("\n"):
                stripped = line.strip()
                if not stripped:
                    continue

                # Heuristic: heading if it matches numbering pattern or ALL CAPS
                is_heading = bool(
                    re.match(r"^\d+(\.\d+){0,3}\s+[A-Z]", stripped) or
                    re.match(r"^[A-Z][A-Z\s\-]{4,}$", stripped) or
                    re.match(r"^Annex\s+[A-Z]", stripped)
                )

                if is_heading and len(stripped) < 100:
                    # Save previous section
                    if current_lines:
                        sections.append((current_heading, " ".join(current_lines)))
                    current_heading = stripped
                    current_lines = []
                else:
                    current_lines.append(stripped)

    if current_lines:
        sections.append((current_heading, " ".join(current_lines)))

    return sections


def _extract_docx_structured(path: Path) -> list[tuple[str, str]]:
    """Extract DOCX using paragraph styles to detect headings."""
    doc = python_docx.Document(str(path))
    sections: list[tuple[str, str]] = []
    current_heading = "Introduction"
    current_paras:   list[str] = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        is_heading = (
            para.style.name.startswith("Heading") or
            re.match(r"^\d+(\.\d+){0,3}\s+[A-Z]", text) or
            re.match(r"^[A-Z][A-Z\s\-]{4,}$", text)
        )

        if is_heading and len(text) < 120:
            if current_paras:
                sections.append((current_heading, " ".join(current_paras)))
            current_heading = text
            current_paras = []
        else:
            current_paras.append(text)

    if current_paras:
        sections.append((current_heading, " ".join(current_paras)))

    return sections


def _split_by_headings(text: str) -> list[tuple[str, str]]:
    """Fallback: split plain text by detected heading patterns."""
    sections: list[tuple[str, str]] = []
    current_heading = "Introduction"
    current_lines:   list[str] = []

    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        if _HEADING_RE.match(stripped) and len(stripped) < 100:
            if current_lines:
                sections.append((current_heading, " ".join(current_lines)))
            current_heading = stripped
            current_lines = []
        else:
            current_lines.append(stripped)

    if current_lines:
        sections.append((current_heading, " ".join(current_lines)))

    return sections


# ─────────────────────────────────────────────────────────────
# STRUCTURE-AWARE CHUNKING
# ─────────────────────────────────────────────────────────────

def chunk_sections(
    sections:   list[tuple[str, str]],
    chunk_size: int = CHUNK_SIZE,
    overlap:    int = CHUNK_OVERLAP,
) -> list[TextChunk]:
    """
    Chunk text while preserving section heading as context.
    Each chunk knows exactly which section it came from — the LLM
    uses this to infer the control domain without guessing.
    """
    chunks: list[TextChunk] = []
    idx = 0

    for heading, text in sections:
        words = text.split()
        if not words:
            continue

        start = 0
        while start < len(words):
            end        = start + chunk_size
            chunk_text = " ".join(words[start:end])

            if len(chunk_text.split()) > 20:   # skip trivially short chunks
                chunks.append(TextChunk(
                    text=chunk_text,
                    section_heading=heading,
                    chunk_index=idx,
                ))
                idx += 1

            start += chunk_size - overlap

    log.info("Created %d structured chunks across %d sections", len(chunks), len(sections))
    return chunks


# ─────────────────────────────────────────────────────────────
# TRIAGE PASS  — cheap binary screen: "any obligations here?"
# ─────────────────────────────────────────────────────────────
#
# WHY A TRIAGE PASS?
#   Standards documents contain large amounts of introductory text,
#   bibliographic references, definitions, and normative references
#   that contain zero actionable obligations.
#   Sending every chunk to a full expert extraction prompt wastes
#   tokens and slows the pipeline. A single-sentence triage prompt
#   costs ~1/10th of the extraction prompt and filters out ~40-60%
#   of chunks (definitions, forewords, bibliography, TOC pages).
#
# The triage prompt is intentionally broad — it passes anything that
# *might* contain an obligation. False negatives (missing a real
# requirement) are much worse than false positives (wasting one
# extraction call on a definitions chunk).

_TRIAGE_SYSTEM = """You are a compliance auditor pre-screening document sections.
Answer YES if the text contains ANY of the following (explicit or implied):
  - Obligations, duties, or responsibilities placed on any actor
  - Actions that must, shall, should, or may be performed
  - Restrictions, prohibitions, or limitations
  - Conditions under which something must happen
  - Accountability or ownership statements
  - Process or procedural requirements
  - Security, privacy, or operational controls (even if described narratively)
Answer NO only if the text is purely definitional, bibliographic, or administrative
(e.g. a table of contents, list of references, or glossary entry with no obligations).
Reply with a single word: YES or NO."""


def triage_chunk(chunk: TextChunk) -> tuple[TextChunk, bool]:
    """Returns (chunk, should_extract)."""
    prompt = (
        f"Section: {chunk.section_heading}\n\n"
        f"Text:\n{chunk.text[:1200]}"
    )
    try:
        resp = azure_client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[
                {"role": "system", "content": _TRIAGE_SYSTEM},
                {"role": "user",   "content": prompt},
            ],
            temperature=0,
            max_tokens=5,
        )
        answer = resp.choices[0].message.content.strip().upper()
        return (chunk, answer.startswith("Y"))
    except Exception:
        log.warning("Triage failed for chunk %d — defaulting to YES", chunk.chunk_index, exc_info=True)
        return (chunk, True)   # fail open: don't drop chunks on error


def triage_all_chunks(chunks: list[TextChunk]) -> list[TextChunk]:
    """Run triage in parallel; return only chunks that passed."""
    passed: list[TextChunk] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(triage_chunk, c): c for c in chunks}
        for fut in as_completed(futures):
            try:
                chunk, should_extract = fut.result()
                if should_extract:
                    passed.append(chunk)
            except Exception:
                c = futures[fut]
                log.error("Triage future raised for chunk %d", c.chunk_index, exc_info=True)
                passed.append(c)   # fail open

    # Restore original order
    passed.sort(key=lambda c: c.chunk_index)
    log.info(
        "Triage: %d / %d chunks passed (%.0f%% filtered)",
        len(passed), len(chunks),
        100 * (1 - len(passed) / max(len(chunks), 1)),
    )
    return passed


# ─────────────────────────────────────────────────────────────
# EXPERT EXTRACTION PASS
# ─────────────────────────────────────────────────────────────
#
# The auditor persona is critical. A keyword extractor asks:
#   "Does this sentence contain the word 'requirement'?"
# A compliance auditor asks:
#   "Does this sentence impose an obligation on anyone, now or conditionally?"
#
# RFC 2119 / ISO modal verb semantics encoded in the prompt:
#   SHALL / MUST         → mandatory obligation
#   SHOULD               → strongly recommended (treated as obligation in audit)
#   MAY                  → optional (still captured — auditors track discretionary controls)
#   SHALL NOT / MUST NOT → prohibition (negative requirement)
#
# Beyond modals, we teach the LLM to detect:
#   - Passive voice obligations: "logs are retained", "access is granted only if"
#   - Responsibility assignments: "the CISO is responsible for"
#   - Implicit requirements from definitions: "an authorized user is one who has..."
#   - Conditional requirements: "where personal data is processed, the controller shall"
#   - Process descriptions that imply ongoing obligations

_EXPERT_SYSTEM = """You are a senior compliance auditor and standards expert with 20 years of experience
auditing against ISO 27001, NIST CSF, SOC 2, PCI-DSS, GDPR, HIPAA, and similar frameworks.

Your task: read a text chunk from a standards document and extract EVERY obligation it contains,
whether explicit or implied. Do not limit yourself to sentences that use the word "requirement".

WHAT TO EXTRACT — any sentence or clause that:
  1. Uses modal verbs: shall, must, should, will, may, is required to, is expected to
  2. Uses passive constructions imposing obligations: "access is restricted", "data is encrypted",
     "logs are retained", "approval must be obtained"
  3. Assigns responsibility or accountability: "the organization is responsible for",
     "system owners shall ensure", "management must approve"
  4. States a prohibition: "shall not", "is prohibited", "is not permitted", "must not"
  5. Describes a condition with a consequence: "where X applies, Y must be implemented"
  6. Implies an ongoing operational requirement through a process description
  7. Sets a measurable threshold: "within 24 hours", "at least annually", "for a minimum of 90 days"

OBLIGATION TYPES — classify each as one of:
  mandatory      → shall / must / is required (non-negotiable)
  recommended    → should / is expected (strong guidance, expected in audit)
  conditional    → "where X, then Y" (applies under specific circumstances)
  negative       → prohibition or restriction (shall not / is prohibited)
  responsibility → accountability or ownership assignment
  operational    → implied by a process description (ongoing obligation)

OUTPUT FORMAT — return ONLY a valid JSON array, no markdown, no prose:
[
  {
    "section_number":  "<clause/section number from heading, e.g. '5.1.2', or '' if not visible>",
    "section_heading": "<the section heading passed to you>",
    "topic":           "<high-level control domain, e.g. 'Access Control', 'Incident Management'>",
    "sub_topic":       "<specific area, e.g. 'Privileged Access', 'Detection and Reporting'>",
    "obligation_type": "<mandatory|recommended|conditional|negative|responsibility|operational>",
    "actor":           "<who bears the obligation, e.g. 'The organization', 'System administrators', 'Data processors'>",
    "condition":       "<if this is a conditional obligation, state the condition; otherwise empty string>",
    "requirement":     "<the obligation restated as a clear, actionable requirement in one sentence>",
    "rationale":       "<one sentence: why this control matters — inferred from context if not stated>"
  }
]

If no obligations exist in the chunk, return [].
Never invent requirements not grounded in the text. Do not paraphrase so loosely that you lose specificity."""


def _repair_json(raw: str) -> str:
    raw = re.sub(r"```(?:json)?", "", raw).replace("```", "").strip()
    start = raw.find("[")
    end   = raw.rfind("]")
    if start != -1 and end != -1:
        return raw[start:end + 1]
    return raw


def extract_from_chunk(standard: str, chunk: TextChunk) -> list[Requirement]:
    """
    Expert auditor extraction for a single triaged chunk.
    The section heading is passed explicitly so the LLM knows its context
    without having to infer it from the text alone.
    """
    prompt = (
        f"Standard being audited: {standard}\n"
        f"Document section: {chunk.section_heading}\n\n"
        f"Text chunk:\n{chunk.text}"
    )

    try:
        resp = azure_client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[
                {"role": "system", "content": _EXPERT_SYSTEM},
                {"role": "user",   "content": prompt},
            ],
            temperature=0,
            max_tokens=2000,
        )
        raw      = resp.choices[0].message.content.strip()
        repaired = _repair_json(raw)

        try:
            items = json.loads(repaired)
        except json.JSONDecodeError:
            log.warning(
                "JSON parse failed for chunk %d. Raw snippet: %.200s",
                chunk.chunk_index, raw,
            )
            return []

        requirements: list[Requirement] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            req_text = item.get("requirement", "").strip()
            if not req_text:
                continue
            requirements.append(Requirement(
                source_name     = standard,
                section_number  = item.get("section_number",  ""),
                section_heading = item.get("section_heading", chunk.section_heading),
                topic           = item.get("topic",           ""),
                sub_topic       = item.get("sub_topic",       ""),
                obligation_type = item.get("obligation_type", "mandatory"),
                actor           = item.get("actor",           ""),
                condition       = item.get("condition",       ""),
                requirement     = req_text,
                rationale       = item.get("rationale",       ""),
            ))

        return requirements

    except Exception:
        log.error("Expert extraction failed for chunk %d", chunk.chunk_index, exc_info=True)
        return []


def extract_all_parallel(standard: str, chunks: list[TextChunk]) -> list[Requirement]:
    """Run expert extraction for all triaged chunks in parallel."""
    all_requirements: list[Requirement] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(extract_from_chunk, standard, c): c for c in chunks}
        for fut in as_completed(futures):
            chunk = futures[fut]
            try:
                result = fut.result()
                all_requirements.extend(result)
                log.debug("Chunk %d → %d requirements", chunk.chunk_index, len(result))
            except Exception:
                log.error("Future raised for chunk %d", chunk.chunk_index, exc_info=True)

    log.info("Raw extraction total: %d requirements", len(all_requirements))
    return all_requirements


# ─────────────────────────────────────────────────────────────
# FUZZY DEDUPLICATION
# ─────────────────────────────────────────────────────────────

def deduplicate(
    requirements: list[Requirement],
    threshold:    int = DEDUP_THRESHOLD,
) -> list[Requirement]:
    """
    Fuzzy dedup using rapidfuzz.partial_ratio.
    Handles paraphrased duplicates that arise from overlapping chunks
    — exact-string match would miss these entirely.
    Preference: keep the version with richer metadata (more fields filled).
    """
    def richness(r: Requirement) -> int:
        return sum([
            bool(r.section_number), bool(r.topic), bool(r.sub_topic),
            bool(r.actor), bool(r.condition), bool(r.rationale),
        ])

    unique: list[Requirement] = []

    for req in requirements:
        text = req.requirement.lower().strip()
        dup_idx = next(
            (i for i, u in enumerate(unique)
             if fuzz.partial_ratio(text, u.requirement.lower().strip()) >= threshold),
            None,
        )
        if dup_idx is None:
            unique.append(req)
        else:
            # Keep the richer of the two
            if richness(req) > richness(unique[dup_idx]):
                unique[dup_idx] = req

    log.info(
        "After dedup: %d / %d requirements kept (threshold=%d%%)",
        len(unique), len(requirements), threshold,
    )
    return unique


# ─────────────────────────────────────────────────────────────
# EXPORT  — JSON + 4-sheet Excel workbook
# ─────────────────────────────────────────────────────────────

from openpyxl import Workbook as _Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from collections import defaultdict

FONT_NAME = "Arial"
HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"

# (light bg, dark text) per obligation type
_OB_COLOURS: dict[str, tuple[str, str]] = {
    "mandatory":      ("FFCCCC", "C00000"),
    "negative":       ("FFE0E0", "FF0000"),
    "conditional":    ("FCE4D6", "C55A11"),
    "responsibility": ("DDEEFF", "1F497D"),
    "recommended":    ("E2EFDA", "375623"),
    "operational":    ("EDE7F6", "4A148C"),
}
_OB_DEFAULT = ("F2F2F2", "000000")

# (json_key, display_name, column_width)
_COLUMN_DEFS = [
    ("source_name",     "Standard",        18),
    ("section_number",  "Section №",       12),
    ("section_heading", "Section Heading", 30),
    ("topic",           "Topic",           22),
    ("sub_topic",       "Sub-Topic",       22),
    ("obligation_type", "Obligation Type", 16),
    ("actor",           "Actor",           24),
    ("condition",       "Condition",       28),
    ("requirement",     "Requirement",     70),
    ("rationale",       "Rationale",       50),
]

_THIN   = Side(style="thin",   color="CCCCCC")
_MED    = Side(style="medium", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HBORDER = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

_OBLIGATION_ORDER = {
    "mandatory": 0, "negative": 1, "conditional": 2,
    "responsibility": 3, "recommended": 4, "operational": 5,
}


def _ob_colours(ob_type: str) -> tuple[str, str]:
    return _OB_COLOURS.get(ob_type.lower().strip(), _OB_DEFAULT)


def _hcell(ws, row: int, col: int, value: str, bg: str = HEADER_BG) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=FONT_NAME, bold=True, color=HEADER_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _HBORDER


def _dcell(ws, row: int, col: int, value,
           wrap: bool = True, bold: bool = False,
           fg: str = "000000", bg: str = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=FONT_NAME, size=10, bold=bold, color=fg)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border    = _BORDER
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)


def _section_sort_key(r) -> tuple:
    """Natural numeric sort on section number for any Requirement or dict."""
    raw = r.section_number if isinstance(r, Requirement) else r.get("section_number", "")
    parts = re.findall(r"\d+", raw)
    return tuple(int(p) for p in parts) if parts else (999,)


# ── Sheet 1: Requirements ─────────────────────────────────────

def _build_requirements_sheet(wb: _Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Requirements"
    ws.sheet_view.showGridLines = False

    for ci, (_, display, width) in enumerate(_COLUMN_DEFS, 1):
        _hcell(ws, 1, ci, display)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 36

    ob_col = next(i + 1 for i, (k, _, _) in enumerate(_COLUMN_DEFS)
                  if k == "obligation_type")

    for ri, req in enumerate(rows, 2):
        ob_type  = req.get("obligation_type", "")
        bg, fg   = _ob_colours(ob_type)

        for ci, (key, _, _) in enumerate(_COLUMN_DEFS, 1):
            value = req.get(key, "")
            if ci == ob_col:
                _dcell(ws, ri, ci, value, bold=True, fg=fg, bg=bg)
            elif key == "requirement":
                _dcell(ws, ri, ci, value, bold=True)
            else:
                _dcell(ws, ri, ci, value)

        ws.row_dimensions[ri].height = 80

        # Subtle alternating row shade on non-obligation columns
        if ri % 2 == 0:
            for ci, (key, _, _) in enumerate(_COLUMN_DEFS, 1):
                if key != "obligation_type":
                    cell = ws.cell(ri, ci)
                    if not cell.fill or cell.fill.patternType is None:
                        cell.fill = PatternFill("solid", fgColor="F8F9FA")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMN_DEFS))}1"


# ── Sheet 2: By Topic ─────────────────────────────────────────

def _build_by_topic_sheet(wb: _Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("By Topic")
    ws.sheet_view.showGridLines = False

    by_topic: dict[str, list[dict]] = defaultdict(list)
    for req in rows:
        by_topic[req.get("topic", "Uncategorised") or "Uncategorised"].append(req)

    cols = [
        ("section_number",  "Section",     12),
        ("obligation_type", "Type",        15),
        ("actor",           "Actor",       22),
        ("requirement",     "Requirement", 80),
        ("rationale",       "Rationale",   45),
    ]

    current_row = 1
    for topic in sorted(by_topic.keys()):
        reqs = by_topic[topic]

        # Topic band
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(cols))
        c = ws.cell(current_row, 1, topic.upper())
        c.font      = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = _HBORDER
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Sub-header
        for ci, (_, display, width) in enumerate(cols, 1):
            _hcell(ws, current_row, ci, display, bg="48768C")
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Rows — sorted by section number
        for req in sorted(reqs, key=_section_sort_key):
            ob_type = req.get("obligation_type", "")
            bg, fg  = _ob_colours(ob_type)
            for ci, (key, _, _) in enumerate(cols, 1):
                value = req.get(key, "")
                if key == "obligation_type":
                    _dcell(ws, current_row, ci, value, bold=True, fg=fg, bg=bg)
                elif key == "requirement":
                    _dcell(ws, current_row, ci, value, bold=True)
                else:
                    _dcell(ws, current_row, ci, value)
            ws.row_dimensions[current_row].height = 75
            current_row += 1

        current_row += 1   # spacer


# ── Sheet 3: Summary ──────────────────────────────────────────

def _build_summary_sheet(wb: _Workbook, rows: list[dict], standard: str) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value     = f"{standard} — Compliance Requirements Summary"
    ws["A1"].font      = Font(name=FONT_NAME, bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value     = f"Total requirements extracted: {len(rows)}"
    ws["A2"].font      = Font(name=FONT_NAME, size=12, italic=True, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # Obligation type counts
    ob_counts: dict[str, int] = defaultdict(int)
    for req in rows:
        ob_counts[req.get("obligation_type", "unknown")] += 1

    ob_order = ["mandatory", "negative", "conditional",
                "responsibility", "recommended", "operational"]
    sorted_ob = sorted(ob_counts.items(),
                       key=lambda x: ob_order.index(x[0]) if x[0] in ob_order else 99)

    for cell, label in [(ws["A4"], "Obligation Type"), (ws["B4"], "Count"), (ws["C4"], "% of Total")]:
        cell.value     = label
        cell.font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _HBORDER

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14

    data_start = 5
    for i, (ob_type, count) in enumerate(sorted_ob):
        r      = data_start + i
        bg, fg = _ob_colours(ob_type)

        ws.cell(r, 1, ob_type.title()).font      = Font(name=FONT_NAME, bold=True, color=fg)
        ws.cell(r, 1).fill                       = PatternFill("solid", fgColor=bg)
        ws.cell(r, 1).alignment                  = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).border                     = _BORDER

        ws.cell(r, 2, count).font                = Font(name=FONT_NAME, size=11, bold=True)
        ws.cell(r, 2).alignment                  = Alignment(horizontal="center")
        ws.cell(r, 2).border                     = _BORDER

        total_row = data_start + len(sorted_ob)
        ws.cell(r, 3, f"=B{r}/B${total_row}").number_format = "0.0%"
        ws.cell(r, 3).font                       = Font(name=FONT_NAME, size=10)
        ws.cell(r, 3).alignment                  = Alignment(horizontal="center")
        ws.cell(r, 3).border                     = _BORDER

    total_row = data_start + len(sorted_ob)
    ws.cell(total_row, 1, "TOTAL").font          = Font(name=FONT_NAME, bold=True)
    ws.cell(total_row, 1).fill                   = PatternFill("solid", fgColor="E0E0E0")
    ws.cell(total_row, 1).border                 = _BORDER
    ws.cell(total_row, 2, f"=SUM(B{data_start}:B{total_row - 1})")
    ws.cell(total_row, 2).font                   = Font(name=FONT_NAME, bold=True)
    ws.cell(total_row, 2).alignment              = Alignment(horizontal="center")
    ws.cell(total_row, 2).border                 = _BORDER

    # Bar chart
    chart = BarChart()
    chart.type   = "bar"
    chart.title  = "Requirements by Obligation Type"
    chart.y_axis.title = "Obligation Type"
    chart.x_axis.title = "Count"
    chart.style  = 10
    chart.width  = 16
    chart.height = 12
    chart.add_data(Reference(ws, min_col=2, max_col=2,
                             min_row=data_start, max_row=total_row - 1))
    chart.set_categories(Reference(ws, min_col=1, max_col=1,
                                   min_row=data_start, max_row=total_row - 1))
    ws.add_chart(chart, "E4")

    # By-topic breakdown table below
    topic_counts: dict[str, int] = defaultdict(int)
    for req in rows:
        topic_counts[req.get("topic", "Uncategorised") or "Uncategorised"] += 1

    t_start = total_row + 3
    for ci, label in enumerate(["Control Domain (Topic)", "Count"], 1):
        ws.cell(t_start, ci, label).font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
        ws.cell(t_start, ci).fill             = PatternFill("solid", fgColor=HEADER_BG)
        ws.cell(t_start, ci).alignment        = Alignment(horizontal="center")
        ws.cell(t_start, ci).border           = _HBORDER

    for i, (topic, count) in enumerate(
        sorted(topic_counts.items(), key=lambda x: -x[1]), 1
    ):
        r = t_start + i
        ws.cell(r, 1, topic).font      = Font(name=FONT_NAME, size=10)
        ws.cell(r, 1).alignment        = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).border           = _BORDER
        ws.cell(r, 1).fill             = PatternFill("solid", fgColor="F0F4FA" if i % 2 == 0 else "FFFFFF")
        ws.cell(r, 2, count).font      = Font(name=FONT_NAME, size=10, bold=True)
        ws.cell(r, 2).alignment        = Alignment(horizontal="center")
        ws.cell(r, 2).border           = _BORDER


# ── Sheet 4: Legend ───────────────────────────────────────────

def _build_legend_sheet(wb: _Workbook) -> None:
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"].value     = "Obligation Type Reference"
    ws["A1"].font      = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    for ci, h in enumerate(["Obligation Type", "Colour Code", "Meaning"], 1):
        _hcell(ws, 2, ci, h)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72

    descriptions = {
        "mandatory":      "Non-negotiable — the organization SHALL or MUST do this.",
        "negative":       "Prohibition — the organization SHALL NOT or MUST NOT do this.",
        "conditional":    "Applies only when a stated condition is met (e.g. 'where encryption is used').",
        "responsibility": "Assigns accountability or ownership to a specific actor or role.",
        "recommended":    "Strongly advised — SHOULD or IS EXPECTED TO; expected in most audits.",
        "operational":    "Implied by a process description; creates an ongoing operational obligation.",
    }

    for ri, (ob_type, desc) in enumerate(descriptions.items(), 3):
        bg, fg = _ob_colours(ob_type)
        ws.cell(ri, 1, ob_type.title()).font      = Font(name=FONT_NAME, bold=True, color=fg)
        ws.cell(ri, 1).fill                       = PatternFill("solid", fgColor=bg)
        ws.cell(ri, 1).alignment                  = Alignment(vertical="center", indent=1)
        ws.cell(ri, 1).border                     = _BORDER
        ws.cell(ri, 2, f"BG #{bg} / Text #{fg}").font = Font(name=FONT_NAME, size=9, color="666666")
        ws.cell(ri, 2).alignment                       = Alignment(horizontal="center", vertical="center")
        ws.cell(ri, 2).border                          = _BORDER
        ws.cell(ri, 3, desc).font                 = Font(name=FONT_NAME, size=10)
        ws.cell(ri, 3).alignment                  = Alignment(wrap_text=True, vertical="center")
        ws.cell(ri, 3).border                     = _BORDER
        ws.row_dimensions[ri].height              = 36


# ── Public entry point ────────────────────────────────────────

def export_results(requirements: list[Requirement], standard: str) -> None:
    if not requirements:
        log.warning("No requirements to export.")
        return

    slug      = re.sub(r"[^\w]", "_", standard.lower())
    json_path = OUTPUT_DIR / f"{slug}_requirements.json"
    xlsx_path = OUTPUT_DIR / f"{slug}_requirements.xlsx"

    # 1. Save JSON
    rows = sorted([vars(r) for r in requirements], key=_section_sort_key)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    log.info("JSON → %s", json_path)

    # 2. Build workbook
    wb = _Workbook()
    _build_requirements_sheet(wb, rows)
    _build_by_topic_sheet(wb, rows)
    _build_summary_sheet(wb, rows, standard)
    _build_legend_sheet(wb)
    wb.save(xlsx_path)
    log.info("Excel → %s  [sheets: %s]", xlsx_path, ", ".join(wb.sheetnames))

    print(f"\n  📄 JSON:  {json_path}")
    print(f"  📊 Excel: {xlsx_path}  ({len(wb.sheetnames)} sheets)")


# ─────────────────────────────────────────────────────────────
# FULL PIPELINE
# ─────────────────────────────────────────────────────────────

def run_pipeline(standard: str) -> PipelineResult:
    result = PipelineResult(standard=standard)
    log.info("═══ Pipeline start: %s ═══", standard)

    # 1. Search + select canonical document
    candidates = search_web(generate_queries(standard))
    if not candidates:
        result.error = "No search results found."
        return result

    candidates = score_candidates(candidates, standard)
    candidates = fetch_previews_parallel(candidates)
    chosen_url = choose_document_llm(standard, candidates)

    if not chosen_url:
        chosen_url = candidates[0].url
        log.warning("LLM selection failed — using top heuristic: %s", chosen_url)

    result.chosen_url = chosen_url
    log.info("Document selected: %s", chosen_url)

    # 2. Download
    local_path = download_file(chosen_url)
    if not local_path:
        result.error = f"Download failed: {chosen_url}"
        return result
    result.local_path = local_path

    # 3. Structure-aware text extraction
    sections = extract_text_with_structure(local_path)
    if not sections:
        result.error = "Could not extract text from document."
        return result
    log.info("Extracted %d sections", len(sections))

    # 4. Chunk (each chunk knows its section heading)
    chunks = chunk_sections(sections)
    if not chunks:
        result.error = "Chunking produced no output."
        return result

    # 5. TRIAGE — filter out non-obligation chunks cheaply
    triaged_chunks = triage_all_chunks(chunks)
    if not triaged_chunks:
        result.error = "Triage filtered out all chunks — document may be non-normative."
        return result

    # 6. EXPERT EXTRACTION — auditor reads every triaged chunk
    raw_requirements = extract_all_parallel(standard, triaged_chunks)

    # 7. Fuzzy dedup
    requirements = deduplicate(raw_requirements)
    result.requirements = requirements

    # 8. Export (sorts internally by section number, then builds 4-sheet workbook)
    export_results(requirements, standard)

    log.info("═══ Done: %d requirements for '%s' ═══", len(requirements), standard)
    return result


# ─────────────────────────────────────────────────────────────
# ENTRYPOINT
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    standard = input("Enter standard/framework name (e.g. 'ISO 27001'): ").strip()

    if not standard:
        print("No standard provided.")
        raise SystemExit(1)

    result = run_pipeline(standard)

    if result.error:
        print(f"\n✗ Pipeline failed: {result.error}")
        raise SystemExit(1)

    # Obligation type breakdown
    from collections import Counter
    breakdown = Counter(r.obligation_type for r in result.requirements)
    print(f"\n✓ {len(result.requirements)} requirements extracted for '{standard}'")
    for ob_type, count in sorted(breakdown.items(), key=lambda x: -x[1]):
        print(f"   {ob_type:<16} {count}")
