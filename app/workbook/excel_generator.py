"""
Excel workbook generator for compliance analysis outputs.
Creates multi-sheet workbooks with formatted compliance data.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path("outputs/workbooks")


class ExcelGenerator:
    """
    Generates professional compliance workbooks using openpyxl.
    Creates multiple sheets for frameworks, requirements, controls, and topics.
    """

    # Color scheme
    COLORS = {
        "header_bg": "1F4E79",      # Dark blue
        "header_fg": "FFFFFF",       # White
        "accent_bg": "2E75B6",       # Medium blue
        "accent_fg": "FFFFFF",
        "alt_row": "D6E4F0",         # Light blue
        "border": "B0C4DE",
        "framework_row": "E8F4FD",
        "control_high": "FDECEA",    # Light red for high risk
        "control_medium": "FFF9E6",  # Light yellow for medium risk
        "control_low": "E8F5E9",     # Light green for low risk
    }

    def __init__(self, output_dir: Path = OUTPUT_DIR):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        frameworks: list[dict],
        records: list[dict],
        controls: list[dict],
        topics: list[dict],
        output_name: str = "compliance_workbook",
    ) -> str:
        """
        Generate a complete compliance Excel workbook.

        Returns:
            Path to the generated workbook file
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment, Border, Font, PatternFill, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed")
            raise

        logger.info("Creating compliance workbook...")
        wb = openpyxl.Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Sheet 1: Framework Overview
        if frameworks:
            self._create_framework_sheet(wb, frameworks, openpyxl)

        # Sheet 2: Document Deconstruction
        if records:
            self._create_records_sheet(wb, records, openpyxl)

        # Sheet 3: Control Library
        if controls:
            self._create_controls_sheet(wb, controls, openpyxl)

        # Sheet 4: Topic/Theme Analysis
        if topics:
            self._create_topics_sheet(wb, topics, openpyxl)

        # Sheet 5: Summary Dashboard
        self._create_summary_sheet(wb, frameworks, records, controls, topics, openpyxl)

        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{output_name}_{timestamp}.xlsx"
        output_path = self.output_dir / filename
        wb.save(str(output_path))
        logger.info(f"Workbook saved: {output_path}")

        return str(output_path)

    def _create_framework_sheet(self, wb, frameworks: list[dict], openpyxl):
        """Sheet 1: Framework Overview."""
        ws = wb.create_sheet("📋 Frameworks")
        ws.sheet_view.showGridLines = True

        headers = ["Framework", "Year", "Authority", "Applicable Industries", "Summary"]
        col_widths = [20, 8, 30, 35, 70]

        self._write_header_row(ws, headers, openpyxl)
        self._set_column_widths(ws, col_widths)

        for i, fw in enumerate(frameworks):
            row = i + 2
            ws.cell(row=row, column=1, value=fw.get("name", ""))
            ws.cell(row=row, column=2, value=fw.get("year", ""))
            ws.cell(row=row, column=3, value=fw.get("authority", ""))
            industries = fw.get("industries", [])
            ws.cell(row=row, column=4, value=", ".join(industries) if isinstance(industries, list) else str(industries))
            ws.cell(row=row, column=5, value=fw.get("summary", ""))

            # Alternate row coloring
            fill_color = self.COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
            self._apply_row_fill(ws, row, 1, 5, fill_color, openpyxl)

        # Word wrap for summary column
        for row in ws.iter_rows(min_row=2, max_row=len(frameworks) + 1, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 30

    def _create_records_sheet(self, wb, records: list[dict], openpyxl):
        """Sheet 2: Document Deconstruction."""
        ws = wb.create_sheet("📄 Requirements")
        headers = ["Source", "Topic", "Subtopic", "Section", "Requirement Text", "Summary", "Theme"]
        col_widths = [15, 20, 20, 12, 60, 40, 20]

        self._write_header_row(ws, headers, openpyxl)
        self._set_column_widths(ws, col_widths)

        for i, rec in enumerate(records[:5000]):  # Cap at 5000 rows
            row = i + 2
            ws.cell(row=row, column=1, value=rec.get("source", ""))
            ws.cell(row=row, column=2, value=rec.get("topic", ""))
            ws.cell(row=row, column=3, value=rec.get("subtopic", ""))
            ws.cell(row=row, column=4, value=rec.get("section_number", ""))
            ws.cell(row=row, column=5, value=rec.get("requirement_text", ""))
            ws.cell(row=row, column=6, value=rec.get("requirement_summary", ""))
            ws.cell(row=row, column=7, value=rec.get("theme", ""))

            fill_color = self.COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
            self._apply_row_fill(ws, row, 1, 7, fill_color, openpyxl)

        # Wrap text in requirement column
        for row in ws.iter_rows(min_row=2, max_row=len(records) + 1, min_col=5, max_col=6):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_controls_sheet(self, wb, controls: list[dict], openpyxl):
        """Sheet 3: Control Library."""
        ws = wb.create_sheet("🛡️ Control Library")
        headers = [
            "Control Theme", "Category", "Subcategory",
            "Control Requirement", "Test Procedure",
            "Risk Narrative", "Mapped Section", "Framework"
        ]
        col_widths = [20, 20, 20, 50, 40, 35, 15, 15]

        self._write_header_row(ws, headers, openpyxl)
        self._set_column_widths(ws, col_widths)

        for i, ctrl in enumerate(controls[:3000]):
            row = i + 2
            ws.cell(row=row, column=1, value=ctrl.get("control_theme", ""))
            ws.cell(row=row, column=2, value=ctrl.get("control_category", ""))
            ws.cell(row=row, column=3, value=ctrl.get("control_subcategory", ""))
            ws.cell(row=row, column=4, value=ctrl.get("control_requirement", ""))
            ws.cell(row=row, column=5, value=ctrl.get("test_procedure", ""))
            ws.cell(row=row, column=6, value=ctrl.get("risk_narrative", ""))
            ws.cell(row=row, column=7, value=ctrl.get("mapped_section", ""))
            ws.cell(row=row, column=8, value=ctrl.get("framework_source", ""))

            fill_color = self.COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
            self._apply_row_fill(ws, row, 1, 8, fill_color, openpyxl)

        # Wrap text in requirement and procedure columns
        for row in ws.iter_rows(min_row=2, max_row=len(controls) + 1, min_col=4, max_col=6):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        ws.auto_filter.ref = ws.dimensions

    def _create_topics_sheet(self, wb, topics: list[dict], openpyxl):
        """Sheet 4: Topic/Theme Analysis."""
        ws = wb.create_sheet("🔍 Themes")
        headers = ["Topic ID", "Theme Label", "Category", "Description", "Keywords", "Associated Sections"]
        col_widths = [10, 25, 20, 50, 40, 30]

        self._write_header_row(ws, headers, openpyxl)
        self._set_column_widths(ws, col_widths)

        for i, topic in enumerate(topics):
            row = i + 2
            ws.cell(row=row, column=1, value=topic.get("topic_id", i))
            ws.cell(row=row, column=2, value=topic.get("label", ""))
            ws.cell(row=row, column=3, value=topic.get("theme_category", ""))
            ws.cell(row=row, column=4, value=topic.get("description", ""))
            keywords = topic.get("keywords", [])
            ws.cell(row=row, column=5, value=", ".join(keywords) if isinstance(keywords, list) else str(keywords))
            sections = topic.get("associated_sections", [])
            ws.cell(row=row, column=6, value=", ".join(str(s) for s in sections) if isinstance(sections, list) else "")

            fill_color = self.COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
            self._apply_row_fill(ws, row, 1, 6, fill_color, openpyxl)

        for row in ws.iter_rows(min_row=2, max_row=len(topics) + 1, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    def _create_summary_sheet(self, wb, frameworks, records, controls, topics, openpyxl):
        """Sheet 5: Summary Dashboard."""
        ws = wb.create_sheet("📊 Summary", 0)  # Insert at beginning

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "🏛️ Regulatory Expert Consultant — Analysis Summary"
        title_cell.font = openpyxl.styles.Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor=self.COLORS["header_bg"])
        title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Generated at
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = openpyxl.styles.Font(italic=True, color="666666")
        ws.merge_cells("A2:F2")

        # Stats
        stats = [
            ("Frameworks Identified", len(frameworks)),
            ("Requirements Extracted", len(records)),
            ("Compliance Controls", len(controls)),
            ("Themes Discovered", len(topics)),
        ]

        ws["A4"] = "Analysis Statistics"
        ws["A4"].font = openpyxl.styles.Font(bold=True, size=13)

        for i, (label, value) in enumerate(stats):
            row = 5 + i
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            ws.cell(row=row, column=2).font = openpyxl.styles.Font(size=12, color="1F4E79")

        # Framework list
        if frameworks:
            ws["A11"] = "Analyzed Frameworks"
            ws["A11"].font = openpyxl.styles.Font(bold=True, size=13)
            for i, fw in enumerate(frameworks):
                row = 12 + i
                ws.cell(row=row, column=1, value=fw.get("name", ""))
                ws.cell(row=row, column=2, value=fw.get("year", ""))
                ws.cell(row=row, column=3, value=fw.get("authority", ""))

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30

    # ─── Helpers ───────────────────────────────────

    def _write_header_row(self, ws, headers: list[str], openpyxl):
        """Write a styled header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(name="Calibri", bold=True, color=self.COLORS["header_fg"])
            cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor=self.COLORS["header_bg"])
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    def _set_column_widths(self, ws, widths: list[int]):
        """Set column widths."""
        from openpyxl.utils import get_column_letter
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _apply_row_fill(self, ws, row: int, min_col: int, max_col: int, color: str, openpyxl):
        """Apply background fill to a row."""
        fill = openpyxl.styles.PatternFill(patternType="solid", fgColor=color)
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).fill = fill
